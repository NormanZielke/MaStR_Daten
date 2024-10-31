import pandas as pd
import numpy as np
from openpyxl import load_workbook

from functions import select_columns, berechne_stilllegung, kraftwerke_nach_cutoff, kraftwerke_still

# runtime of technologies
betriebsdauer = {
    'Verbrennungsmotor': 30,
    'Gasturbinen mit Abhitzekessel': 40,
    'Stirlingmotor': 25,
    'Brennstoffzelle': 15,
    'Gegendruckmaschine mit Entnahme': 50,
    'Gasturbinen ohne Abhitzekessel': 30,
    'Kondensationsmaschine mit Entnahme':60
}



cutoff_date = pd.to_datetime('2045-01-01')

#---------------------------------------------------------------------------------------------------------------------->

# Kiel

df_comb_Kiel = pd.read_csv("regions_csv/kiel_combustion.csv",
                        sep=",")
# filter MaStR by selected columns
Kiel_combustion = select_columns(df_comb_Kiel)
# calculate expire date by "betriebsdauer" and gives power plants, which are still in progress
Kiel_combustion_nach_2045 = kraftwerke_nach_cutoff(Kiel_combustion,cutoff_date,betriebsdauer)
# gives power plants, which are expired until cutoff_date
Kiel_combustion_bis_2045_sill = kraftwerke_still(Kiel_combustion,Kiel_combustion_nach_2045)

"""
df = Kiel_combustion[Kiel_combustion["Strasse"] == "Am Kiel-Kanal 2"]

df = Kiel_combustion[Kiel_combustion["Postleitzahl"] == 24106]

df = Kiel_combustion[Kiel_combustion["ThermischeNutzleistung"] == 390]
df = Kiel_combustion[Kiel_combustion["ElektrischeKwkLeistung"] == 330]
"""

#---------------------------------------------------------------------------------------------------------------------->

# Rüdersdorf bei Berlin

df_comb_Ruedersdorf = pd.read_csv("regions_csv/Ruedersdorf_combustion.csv",
                        sep=",")

Ruedersdorf_combustion = select_columns(df_comb_Ruedersdorf)

Ruedersdorf_combustion_nach_2045 = kraftwerke_nach_cutoff(Ruedersdorf_combustion,cutoff_date,betriebsdauer)

Ruedersdorf_combustion_bis_2045_still = kraftwerke_still(Ruedersdorf_combustion,Ruedersdorf_combustion_nach_2045)

#---------------------------------------------------------------------------------------------------------------------->

# Erkner

df_comb_Erkner = pd.read_csv("regions_csv/Erkner_combustion.csv",
                        sep=",")

Erkner_combustion = select_columns(df_comb_Erkner)

Erkner_combustion_nach_2045 = kraftwerke_nach_cutoff(Erkner_combustion,cutoff_date,betriebsdauer)

Erkner_combustion_bis_2045_still = kraftwerke_still(Erkner_combustion,Erkner_combustion_nach_2045)

#---------------------------------------------------------------------------------------------------------------------->

# Grünheide Mark

df_comb_GruenH = pd.read_csv("regions_csv/GruenH_combustion.csv",
                        sep=",")

GruenH_combustion = select_columns(df_comb_GruenH)

GruenH_combustion_nach_2045 = kraftwerke_nach_cutoff(GruenH_combustion,cutoff_date,betriebsdauer)

GruenH_combustion_bis_2045_still = kraftwerke_still(GruenH_combustion,GruenH_combustion_nach_2045)

#---------------------------------------------------------------------------------------------------------------------->

# Strausberg

df_comb_Strausberg = pd.read_csv("regions_csv/Strausberg_combustion.csv",
                        sep=",")

Strausberg_combustion = select_columns(df_comb_Strausberg)

Strausberg_combustion_nach_2045 = kraftwerke_nach_cutoff(Strausberg_combustion,cutoff_date,betriebsdauer)

Strausberg_combustion_bis_2045_still = kraftwerke_still(Strausberg_combustion,Strausberg_combustion_nach_2045)

#---------------------------------------------------------------------------------------------------------------------->

# save data

with pd.ExcelWriter("SLE_combustion_portfolio.xlsx") as writer:
    Kiel_combustion.to_excel(writer, sheet_name="Kiel", index=False)
    Ruedersdorf_combustion.to_excel(writer, sheet_name="Ruedersdorf", index=False)
    Erkner_combustion.to_excel(writer, sheet_name="Erkner", index=False)
    GruenH_combustion.to_excel(writer, sheet_name="GruenHeide", index=False)
    Strausberg_combustion.to_excel(writer, sheet_name="Strausberg", index=False)

# Lade die erstellte Excel-Datei mit openpyxl
wb = load_workbook("SLE_combustion_portfolio.xlsx")

# Funktion zum Anpassen der Spaltenbreite
def adjust_column_width(sheet):
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

# Passe die Spaltenbreiten für jedes Blatt an
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    adjust_column_width(sheet)

# Speichere die Datei mit den angepassten Spaltenbreiten
wb.save("SLE_combustion_portfolio.xlsx")

#---------------------------------------------------------------------------------------------------------------------->





